#encoding=utf-8
import openpyxl
from openpyxl.styles import Border,Side,Font
import time

class ParseExcel(object):
    def __init__(self):
        self.workbook=None
        self.excelFile=None
        self.font=Font(color=None)
        self.colorDict={'red':'FFFF3030','green':'FF008B00'}

    def loadWorkBook(self,excelPathAndName):
        #将excel文件加载到内存，并获取其workbook对象
        try:
            self.workbook=openpyxl.load_workbook(excelPathAndName)
        except Exception,e:
            raise e
        self.excelFile=excelPathAndName
        return self.workbook
    def getSheetByName(self,sheetName):
        try:
            sheet=self.workbook[sheetName]
            return sheet
        except Exception,e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        try:
            sheetName=self.workbook.sheetnames[sheetIndex]
        except Exception,e:
            raise e
        sheet=self.workbook[sheetName]
        return sheet

    def getTotalRowsNumber(self,sheet):
        return sheet.max_row

    def getTotalColsNumber(self,sheet):
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        return sheet.min_row

    def getStartColNumber(self,sheet):
        return sheet.min_column

    def getSingleRow(self,sheet,rowNo):
        #获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple
        #下标从1开始，sheet.rows[1]表示第一行
        try:
            return list(sheet.rows)[rowNo-1]
        except Exception,e:
            raise e

    def getSingleColumn(self,sheet,colNo):
        #获取sheet中某一列，返回的是这一列所有的数据组成的tuple
        #下标从1开始，sheet.columns[1]表示第一列
        try:
            return list(sheet.columns)[colNo-1]
        except Exception,e:
            raise e

    def getValueInCell(self,sheet,coordinate=None,rowNo=None,colNo=None):
        #根据单元格所在的位置索引获取该单元格中的值，下标从1开始
        #sheet.cell(row=1,column=1).value,表示excel中第一行第一列的值
        if coordinate != None:#coordinate指坐标，如['A2']
            try:
                return sheet[coordinate].value
            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colNo).value
            except Exception,e:
                raise e
        else:
            raise Exception("Argument exception! ")

    def writeCell(self,sheet,content,coordinate=None,rowNo=None,colNo=None,color=None):
        #根据单元格在excel中的行、列号，或坐标值向单元格中写入数据
        #color标识字体的颜色的名字，如red，green
        if coordinate:
            try:
                sheet[coordinate].value=content
                if color:
                    sheet[coordinate].font=Font(color=self.colorDict[color])
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        elif coordinate == None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo,column=colNo).value = content
                if color:
                    sheet.cell(row=rowNo,column=color).font=Font(color=self.colorDict[color])
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        else:
            raise Exception("Argument exception!")

    def writeCurrentTimeInCell(self,sheet,coordinate=None,rowNo=None,colNo=None):
        #写入当前时间，下标从1开始，如['A1']
        timeArray=time.localtime(time.time())
        currentTime=time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet[coordinate].value=currentTime
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        elif coordinate == None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row == rowNo,column=colNo).value=currentTime
                self.workbook.save(self.excelFile)
            except Exception,e:
                raise e
        else:
            raise Exception("Argument exception!")

if __name__=='__main__':
    pe=ParseExcel()
    pe.loadWorkBook(r'd:\\testdata.xlsx')
    sheetObj=pe.getSheetByName(u"API")
    print u"用index号获取sheet对象的名字：",pe.getSheetByIndex(0)
    sheet=pe.getSheetByIndex(1)
    print type(sheet)
    print pe.getTotalRowsNumber(sheetObj)
    print pe.getTotalColsNumber(sheetObj)
    print pe.getTotalRowsNumber(sheet)
    print pe.getTotalColsNumber(sheet)
    #print list(sheet.rows)[1]
    rows=pe.getSingleRow(sheet,1)
    # for i in rows:
    #     if i.value:
    #         print i.value
    print pe.getValueInCell(sheetObj,'A1')
    pe.writeCell(sheet,"xia","A2")
    print pe.getValueInCell(sheetObj,"A2")
    pe.writeCurrentTimeInCell(sheetObj,"A3")





























